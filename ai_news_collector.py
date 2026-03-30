#!/usr/bin/env python3
"""
AI资讯收集器 v3.0 - 优化版
自动收集AI行业资讯，分类、标记重点、生成报告

优化内容：
- 改进去重算法：标题相似度检测、URL标准化
- 翻译优化：重试机制、超时保护
- 错误恢复：网络容错、请求限流
- HTML优先级：语音 > 视频/图像 > LLM > 大厂 > 其他
"""

import os
import sys
import csv
import re
import json
import time
import argparse
import logging
import hashlib
from pathlib import Path
from datetime import datetime, timedelta
from difflib import SequenceMatcher
from urllib.parse import urlparse, urlunparse

import requests
import feedparser

try:
    from deep_translator import GoogleTranslator, MyMemoryTranslator
    HAS_TRANSLATOR = True
except ImportError:
    HAS_TRANSLATOR = False

try:
    import yaml
    HAS_YAML = True
except ImportError:
    HAS_YAML = False

# ========== 路径工具 ==========
SCRIPT_DIR = Path(__file__).resolve().parent


def find_config(config_path=None):
    """定位config.yaml：优先参数，再找脚本同级，最后找exe同级"""
    if config_path:
        p = Path(config_path)
        if p.exists():
            return p
        raise FileNotFoundError(f"配置文件不存在: {p}")
    for candidate in [SCRIPT_DIR / "config.yaml", Path(sys.executable).parent / "config.yaml"]:
        if candidate.exists():
            return candidate
    return None


# ========== 配置 ==========
def load_config(config_path=None):
    """加载配置文件"""
    cfg_file = find_config(config_path)
    if cfg_file and HAS_YAML:
        with open(cfg_file, encoding="utf-8") as f:
            return yaml.safe_load(f)
    # 回退默认配置
    return {
        "translation": {
            "enabled": True,
            "max_text_length": 600,
            "cache_file": ".translation_cache.json",
            "retry_times": 3,
            "timeout_seconds": 10
        },
        "sources": [
            {"name": "OpenAI Blog", "url": "https://openai.com/blog/rss.xml", "type": "英文", "limit": 15},
            {"name": "量子位", "url": "https://www.qbitai.com/feed", "type": "中文", "limit": 20},
        ],
        "schedule": {"interval_hours": 6, "first_run_delay_minutes": 0},
        "output": {"csv": "ai_news_report.csv", "html": "ai_news_report.html", "list": "ai_news_list.txt", "html_max_items": 100},
        "deduplication": {
            "title_similarity_threshold": 0.85,
            "enable_url_dedup": True,
            "enable_content_dedup": True
        },
        "html_priority": {
            "sections": [
                {"name": "语音技术", "categories": ["语音技术", "TTS", "ASR"], "color": "#22c55e", "companies": []},
                {"name": "视频生成", "categories": ["视频生成"], "color": "#f59e0b", "companies": []},
            ]
        }
    }


# ========== 关键词定义 ==========
AI_KEYWORDS = [
    "ai", "人工智能", "machine learning", "机器学习", "deep learning", "深度学习",
    "llm", "大模型", "foundation model", "基础模型",
    "gpt", "claude", "llama", "gemini", "deepseek", "chatgpt",
    "tts", "语音合成", "text-to-speech", "asr", "语音识别", "whisper",
    "speech", "语音技术", "语音交互", "voice", "audio",
    "video generation", "文生视频", "sora", "vidu", "可灵", "kling",
    "图生视频", "image to video", "runway", "pika", "视频生成",
    "image generation", "图像生成", "stable diffusion", "midjourney", "dall-e",
    "开源", "open source", "hugging face", "github",
    "模型", "算法", "训练", "推理", "参数", "benchmark", "评测", "leaderboard",
    "openai", "google ai", "deepmind", "anthropic", "字节", "百度", "腾讯", "阿里",
    "科大讯飞", "快手", "智谱", "百川", "月之暗面", "minimax",
    "fish audio", "elevenlabs",
    "paper", "arxiv", "论文",
]

CATEGORY_MAP = {
    "语音技术": ["tts", "语音合成", "asr", "语音识别", "whisper", "speech", "语音交互", "fish audio", "elevenlabs", "audio"],
    "TTS": ["tts", "语音合成", "text-to-speech"],
    "ASR": ["asr", "语音识别", "speech recognition"],
    "视频生成": ["video generation", "文生视频", "sora", "vidu", "可灵", "kling", "图生视频", "runway", "pika"],
    "图像生成": ["image generation", "图像生成", "stable diffusion", "midjourney", "dall-e", "文生图"],
    "大模型": ["llm", "大模型", "gpt", "claude", "llama", "gemini", "deepseek", "chatgpt", "多模态模型", "multimodal"],
    "开源发布": ["开源", "open source", "hugging face"],
    "论文研究": ["paper", "arxiv", "论文"],
    "产品发布": ["发布", "上线", "launch", "release", "update", "更新"],
    "融资并购": ["融资", "收购", "投资", "funding", "acquisition", "raise"],
    "技术突破": ["突破", "超越", "sota", "首个", "state-of-the-art"],
    "评测榜单": ["benchmark", "leaderboard", "评测", "排行榜"],
}

COMPANY_MAP = {
    "OpenAI": ["openai", "chatgpt", "gpt-4", "gpt-5", "sora"],
    "Google": ["google", "deepmind", "gemini", "bard"],
    "Meta": ["meta", "facebook", "llama"],
    "Microsoft": ["microsoft", "azure", "copilot"],
    "Anthropic": ["anthropic", "claude"],
    "xAI": ["xai", "grok"],
    "Stability AI": ["stability ai", "stable diffusion", "stability"],
    "Midjourney": ["midjourney"],
    "字节跳动": ["字节", "字节跳动", "抖音", "豆包", "seedance", "doubao"],
    "阿里巴巴": ["阿里", "阿里巴巴", "通义", "千问", "qwen"],
    "腾讯": ["腾讯", "混元", "hunyuan"],
    "百度": ["百度", "文心", "wenxin"],
    "科大讯飞": ["科大讯飞", "讯飞", "spark"],
    "快手": ["快手", "可灵", "kling"],
    "智谱AI": ["智谱", "glm", "chatglm", "zhipu"],
    "月之暗面": ["月之暗面", "moonshot", "kimi"],
    "MiniMax": ["minimax", "abacus", "海螺"],
    "DeepSeek": ["deepseek", "深度求索"],
    "Hugging Face": ["hugging face", "huggingface"],
    "ElevenLabs": ["elevenlabs"],
    "Fish Audio": ["fish audio", "fishaudio"],
    "Runway": ["runway"],
    "Pika": ["pika"],
}

IMPORTANT_KEYWORDS = ["开源", "发布", "融资", "突破", "上线", "收购", "launch", "release", "open source", "funding", "raise", "acquisition"]


# ========== 去重优化 ==========
def normalize_url(url):
    """标准化URL，去除查询参数"""
    if not url:
        return ""
    try:
        parsed = urlparse(url)
        # 保留基本路径，移除查询参数和片段
        normalized = urlunparse((
            parsed.scheme,
            parsed.netloc,
            parsed.path,
            '',  # params
            '',  # query
            ''   # fragment
        ))
        return normalized.lower()
    except Exception:
        return url.lower()


def calculate_similarity(s1, s2):
    """计算两个字符串的相似度（Jaccard）"""
    if not s1 or not s2:
        return 0
    set1 = set(s1.lower())
    set2 = set(s2.lower())
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    return intersection / union if union > 0 else 0


def is_duplicate_title(new_title, existing_titles, threshold=0.85):
    """检查标题是否重复（基于相似度）"""
    new_title_clean = re.sub(r'[^\w\s]', '', new_title.lower()).strip()
    
    for existing_title in existing_titles:
        existing_clean = re.sub(r'[^\w\s]', '', existing_title.lower()).strip()
        similarity = calculate_similarity(new_title_clean, existing_clean)
        if similarity >= threshold:
            return True
    return False


# ========== 翻译优化 ==========
_TRANSLATION_CACHE = {}


def _is_english(text):
    """判断文本是否主要为英文"""
    if not text or len(text) < 5:
        return False
    alpha = sum(1 for c in text if c.isascii() and c.isalpha())
    return alpha / max(len(text), 1) > 0.5


def translate_text(text, cfg=None):
    """翻译英文→中文，带缓存、重试和超时保护"""
    if cfg is None:
        cfg = {}
    if not cfg.get("enabled", True) or not text or not _is_english(text):
        return text

    max_len = cfg.get("max_text_length", 600)
    retry_times = cfg.get("retry_times", 3)
    timeout = cfg.get("timeout_seconds", 10)
    
    text_s = text[:max_len].strip()
    key = hashlib.md5(text_s.encode()).hexdigest()
    if key in _TRANSLATION_CACHE:
        return _TRANSLATION_CACHE[key]

    result = text  # 回退

    if not HAS_TRANSLATOR:
        return result

    # 尝试翻译，带重试
    for attempt in range(retry_times):
        try:
            # 方案1: Google Translate (deep-translator) 无限制
            translator = GoogleTranslator(source="en", target="zh-CN")
            t = translator.translate(text_s)
            if t and t.strip() and "WARNING" not in t and t != text_s:
                result = t
                _TRANSLATION_CACHE[key] = result
                return result
        except Exception as e:
            if attempt < retry_times - 1:
                time.sleep(1)  # 重试前等待
                continue
            logging.debug(f"翻译失败（尝试{attempt+1}次）: {e}")

    # 方案2: MyMemory (每天5000字符免费) - 只尝试一次
    if retry_times == 0:
        try:
            t = MyMemoryTranslator(source="en-CN", target="zh-CN").translate(text_s[:500])
            if t and t.strip() and "WARNING" not in t and "YOU USED" not in t:
                result = t
        except Exception:
            pass

    _TRANSLATION_CACHE[key] = result
    return result


def load_translation_cache(cfg):
    cache_file = cfg.get("cache_file", ".translation_cache.json")
    try:
        with open(cache_file, encoding="utf-8") as f:
            _TRANSLATION_CACHE.update(json.load(f))
    except (FileNotFoundError, json.JSONDecodeError):
        pass


def save_translation_cache(cfg):
    cache_file = cfg.get("cache_file", ".translation_cache.json")
    try:
        with open(cache_file, "w", encoding="utf-8") as f:
            json.dump(_TRANSLATION_CACHE, f, ensure_ascii=False)
    except Exception as e:
        logging.warning(f"保存翻译缓存失败: {e}")


# ========== 文本处理 ==========
def clean_text(text, max_len=200):
    if not text:
        return ""
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) > max_len:
        text = text[:max_len] + "..."
    return text


def generate_key_points(text, max_points=3):
    """从摘要中提取结构化要点"""
    if not text:
        return ""
    text = clean_text(text, 500)
    if not text:
        return ""
    # 按句号/分号/换行分割
    sentences = re.split(r"[。！？；\n]", text)
    sentences = [s.strip() for s in sentences if len(s.strip()) > 6]
    if sentences:
        points = []
        for i, s in enumerate(sentences[:max_points], 1):
            # 去掉开头的冗余词
            s = re.sub(r'^(据悉|据了解|目前|最近|据介绍|对此|记者了解到|消息显示)[，,]?', '', s)
            if s:
                points.append(f"{i}. {s}")
        return "\n".join(points) if points else clean_text(text, 150)
    return clean_text(text, 150)


def is_important(title, summary):
    text = (title + " " + summary).lower()
    return any(kw.lower() in text for kw in IMPORTANT_KEYWORDS)


def smart_category(title, summary):
    """智能分类，使用关键词权重匹配"""
    text = (title + " " + summary).lower()
    
    # 定义分类规则（关键词 + 权重）
    CATEGORY_RULES = [
        # 语音技术相关
        {
            "name": "语音技术",
            "keywords": ["tts", "语音合成", "text-to-speech", "asr", "语音识别", "whisper", "speech", "语音交互", "voice", "fish audio", "elevenlabs", "audio", "情感识别", "情感智能", "对话"],
            "exclude": ["视频", "图像", "视觉", "image", "video", "vision", "音乐", "music", "效果测试", "效果评估", "效果评测", "音质测试", "音质评估", "音质评测", "声学测试", "声学评估", "质量评估", "质量评测", "评测框架", "测试框架", "评估方案", "评测方案", "测试方案", "评估方法", "评测方法", "测试方法", "评估指标", "评测指标", "测试指标"],
            "weight": 3
        },
        # 视觉相关
        {
            "name": "视觉相关",
            "keywords": ["video generation", "文生视频", "sora", "vidu", "可灵", "kling", "图生视频", "image to video", "runway", "pika", "视频生成", "image generation", "图像生成", "stable diffusion", "midjourney", "dall-e", "文生图", "视觉", "vision", "computer vision", "nano banana"],
            "exclude": [],
            "weight": 5  # 提高视觉相关权重，因为"安全"会误判
        },
        # 大模型相关
        {
            "name": "大模型",
            "keywords": ["llm", "大模型", "gpt", "claude", "llama", "gemini", "deepseek", "chatgpt", "多模态模型", "multimodal", "foundation model", "基础模型", "transformer", "bert", "模型发布", "模型更新", "模型能力"],
            "exclude": ["音乐", "audio", "music", "语音", "speech"],  # 排除音乐生成
            "weight": 2
        },
        # 音频/音乐（单独分类，不归入大模型）
        {
            "name": "音频/音乐",
            "keywords": ["音乐", "music", "audio generation", "音乐生成", "音频生成", "lyria", "suno", "udio", "musicai", "作曲", "编曲"],
            "exclude": [],
            "weight": 4  # 提高权重，优先级高于大模型
        },
        # AI安全
        {
            "name": "AI安全",
            "keywords": ["安全门", "风险管控", "safety", "security", "伦理", "alignment", "对齐", "risk", "threat", "srm"],
            "exclude": ["sora", "video", "视觉", "image"],  # 排除视觉相关的内容
            "weight": 3
        },
        # AI应用/产业
        {
            "name": "AI应用",
            "keywords": ["应用", "落地", "场景", "行业", "转型", "赋能", "数字化", "提效", "workflow", "工具", "助手", "培训", "计划", "人才", "中心", "联邦", "公司"],
            "exclude": [],
            "weight": 2
        },
        # 评测榜单
        {
            "name": "评测榜单",
            "keywords": ["benchmark", "leaderboard", "排行榜", "排名", "对比"],
            "exclude": ["效果测试", "效果评估", "效果评测", "音质测试", "音质评估", "音质评测", "声学测试", "声学评估", "质量评估", "质量评测", "评测框架", "测试框架", "评估方案", "评测方案", "测试方案", "评估方法", "评测方法", "测试方法", "评估指标", "评测指标", "测试指标", "质量指标", "效果指标", "准确率效果"],
            "weight": 3
        },
        # 效果测试
        {
            "name": "效果测试",
            "keywords": ["效果测试", "效果评估", "效果评测", "音质测试", "音质评估", "语音效果", "音频效果", "声音测试", "声音评估", "音质评测", "语音质量", "音频质量", "质量评估", "质量评测", "声学测试", "声学评估", "语音合成评测", "asr评测", "tts评测", "语音识别评测", "评估框架", "评测框架", "测试框架", "评估方案", "评测方案", "测试方案", "评估方法", "评测方法", "测试方法", "评估指标", "评测指标", "测试指标", "评估体系", "评测体系", "测试体系", "主观测试", "客观测试", "mos测试", "mean opinion score", "语音质量评测", "音频质量评测", "音质对比", "效果对比", "性能对比", "声学性能", "语音性能", "音频性能", "质量指标", "效果指标", "音质", "声学", "准确率效果"],
            "exclude": [],
            "weight": 4  # 提高权重，优先级高于语音技术和评测榜单
        },
    ]
    
    # 计算每个分类的得分
    category_scores = {}
    for rule in CATEGORY_RULES:
        score = 0
        
        # 检查关键词匹配
        for kw in rule["keywords"]:
            if kw.lower() in text:
                # 检查是否在标题中（权重更高）
                if kw.lower() in title.lower():
                    score += rule["weight"] * 2
                else:
                    score += rule["weight"]
        
        # 检查排除关键词
        for exclude_kw in rule["exclude"]:
            if exclude_kw.lower() in text:
                score = -100  # 严重降权
                break
        
        if score > 0:
            category_scores[rule["name"]] = score
    
    # 返回得分最高的分类
    if category_scores:
        best_category = max(category_scores.items(), key=lambda x: x[1])
        return best_category[0]
    
    # 如果没有匹配，使用传统方法
    for cat, keywords in CATEGORY_MAP.items():
        if any(kw.lower() in text for kw in keywords):
            return cat
    
    return "AI相关"


def extract_company(title, summary):
    text = (title + " " + summary).lower()
    for company, keywords in COMPANY_MAP.items():
        if any(kw.lower() in text for kw in keywords):
            return company
    return "其他"


def contains_ai_keyword(title, summary):
    text = (title + " " + summary).lower()
    return any(kw.lower() in text for kw in AI_KEYWORDS)


# ========== RSS抓取（优化） ==========
def parse_rss_date(published_str):
    """解析RSS发布时间，返回datetime对象"""
    if not published_str:
        return None
    try:
        # feedparser会自动解析published字段为published_parsed
        return datetime.fromtimestamp(time.mktime(feedparser.parse(published_str).entries[0].published_parsed))
    except:
        pass
    
    # 尝试其他常见格式
    for fmt in ["%a, %d %b %Y %H:%M:%S %z", "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"]:
        try:
            return datetime.strptime(published_str, fmt)
        except:
            continue
    return None


def fetch_rss(url, limit=15, timeout=30, days_limit=3):
    """抓取RSS，带超时和错误处理，只返回最近N天的资讯"""
    try:
        # 添加User-Agent，避免被反爬
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        # feedparser的parse不支持timeout参数，使用requests先获取内容
        response = requests.get(url, headers=headers, timeout=timeout)
        response.raise_for_status()
        # 解析RSS内容
        feed = feedparser.parse(response.content)
        
        # 计算时间阈值：最近N天
        cutoff_date = datetime.now() - timedelta(days=days_limit)
        
        items = []
        for entry in feed.entries[:limit * 2]:  # 多抓取一些，过滤后可能不足
            # 提取发布时间
            pub_date_str = getattr(entry, "published", "")
            pub_date = None
            
            # 优先使用published_parsed（feedparser已解析的时间戳）
            if hasattr(entry, "published_parsed") and entry.published_parsed:
                try:
                    pub_date = datetime.fromtimestamp(time.mktime(entry.published_parsed))
                except:
                    pass
            
            # 如果没有parsed，尝试解析字符串
            if not pub_date and pub_date_str:
                pub_date = parse_rss_date(pub_date_str)
            
            # 如果仍然没有时间，使用当前时间
            if not pub_date:
                pub_date = datetime.now()
            
            # 只保留最近N天的资讯
            if pub_date >= cutoff_date:
                items.append({
                    "title": getattr(entry, "title", ""),
                    "link": getattr(entry, "link", ""),
                    "summary": getattr(entry, "summary", "")[:500],
                    "published": pub_date_str,  # 原始字符串，用于显示
                    "published_parsed": pub_date,  # 解析后的时间对象，用于过滤和排序
                })
            
            # 如果已经收集够了，就停止
            if len(items) >= limit:
                break
        
        return items
    except Exception as e:
        logging.warning(f"RSS抓取失败 {url}: {e}")
        return []


# ========== HTML报告（优先级优化） ==========
_CSS = """
:root{--bg:#0d1117;--card:#161b22;--border:#30363d;--text:#e6edf3;--muted:#8b949e;
--green:#3fb950;--blue:#58a6ff;--orange:#d29922;--red:#f85149;--purple:#bc8cff;
--imp-border:#f85149;--imp-bg:rgba(248,81,73,.08)}
*{margin:0;padding:0;box-sizing:border-box}
body{background:var(--bg);color:var(--text);font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',system-ui,sans-serif;
line-height:1.6;padding:32px;max-width:1400px;margin:0 auto}
.header{text-align:center;margin-bottom:36px;padding:32px;background:var(--card);border:1px solid var(--border);border-radius:12px}
.header h1{font-size:28px;margin-bottom:8px}.header h1 span{color:var(--blue)}
.header .meta{color:var(--muted);font-size:14px}
.stats{display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:12px;margin-bottom:36px}
.stat{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:16px;text-align:center}
.stat .num{font-size:26px;font-weight:700;color:var(--green)}.stat .label{font-size:12px;color:var(--muted);margin-top:4px}
section{margin-bottom:32px}section h2{font-size:20px;margin-bottom:16px;padding-bottom:8px;border-bottom:1px solid var(--border)}
.item{background:var(--card);border:1px solid var(--border);border-radius:8px;padding:16px;margin-bottom:10px;
transition:border-color .2s}.item:hover{border-color:var(--blue)}
.item.important{border-left:3px solid var(--imp-border);background:var(--imp-bg)}
.item .title{font-size:15px;font-weight:600;margin-bottom:6px}
.item .meta2{font-size:12px;color:var(--muted);margin-bottom:8px;display:flex;gap:12px;flex-wrap:wrap}
.item .meta2 .tag{padding:2px 8px;border-radius:4px;font-size:11px;font-weight:500}
.tag-cat{background:rgba(88,166,255,.12);color:var(--blue)}
.tag-co{background:rgba(188,140,255,.12);color:var(--purple)}
.tag-imp{background:rgba(248,81,73,.12);color:var(--red)}
.item .points{font-size:13px;color:var(--muted);margin-top:8px;white-space:pre-line}
.item a{color:var(--blue);text-decoration:none;font-size:12px}a:hover{text-decoration:underline}
.footer{text-align:center;color:var(--muted);font-size:12px;padding:20px;border-top:1px solid var(--border);margin-top:40px}
.priority-section{background:var(--card);border:1px solid var(--border);border-left:4px solid var(--blue);border-radius:10px;padding:20px;margin-bottom:20px}
.priority-section h2{font-size:18px;margin-bottom:12px;display:flex;align-items:center;gap:8px}
.priority-section .count{background:rgba(255,255,255,.08);padding:2px 10px;border-radius:12px;font-size:13px;color:var(--muted);font-weight:400}
.other-cat{color:var(--muted);margin:16px 0 8px;font-size:15px;padding-left:8px;border-left:2px solid var(--border)}
.priority-section .item{border-color:transparent;background:transparent;margin-bottom:8px}
.priority-section .item:hover{border-color:var(--border)}

/* 重点领域横向布局 */
.domain-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(300px,1fr));gap:16px;margin-bottom:32px}
.domain-card{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:16px}
.domain-card h3{font-size:16px;margin-bottom:12px;display:flex;align-items:center;gap:8px}
.domain-card .count{background:rgba(255,255,255,.08);padding:2px 8px;border-radius:12px;font-size:12px;color:var(--muted)}
.domain-card .items{margin-top:12px}
.domain-card .item{border-color:transparent;background:transparent;padding:8px;margin-bottom:6px}
.domain-card .item .title{font-size:13px}

/* 大厂动态横向布局 */
.company-section{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:20px;margin-bottom:32px}
.company-section h2{font-size:18px;margin-bottom:20px}
.company-group{margin-bottom:24px}
.company-group h3{font-size:15px;color:var(--muted);margin-bottom:12px;padding-bottom:8px;border-bottom:1px solid var(--border)}
.company-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:12px}
.company-item{background:rgba(255,255,255,.03);border:1px solid var(--border);border-radius:8px;padding:12px}
.company-item .company-name{font-size:13px;font-weight:600;color:var(--blue);margin-bottom:4px}
.company-item .item{border:none;padding:0;margin:0}
.company-item .item .title{font-size:12px;margin-bottom:4px}
.company-item .item .meta2{font-size:11px}

/* AI提效资讯 */
.efficiency-section{background:var(--card);border:1px solid var(--border);border-left:4px solid var(--orange);border-radius:10px;padding:20px;margin-bottom:32px}
.efficiency-section h2{font-size:18px;margin-bottom:16px;color:var(--orange)}

/* 重要资讯速览列表 */
.important-list{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:20px;margin-bottom:32px}
.important-list h2{font-size:18px;margin-bottom:16px;display:flex;align-items:center;gap:8px}
.important-item{display:flex;gap:12px;padding:12px;border-bottom:1px solid var(--border);align-items:flex-start}
.important-item:last-child{border-bottom:none}
.important-item .marker{color:var(--red);font-size:16px}
.important-item .content{flex:1}
.important-item .title{font-size:14px;font-weight:600;margin-bottom:4px}
.important-item .meta{font-size:12px;color:var(--muted);margin-bottom:4px}
.important-item .link{color:var(--blue);text-decoration:none;font-size:12px}
.important-item .link:hover{text-decoration:underline}

/* 筛选器 */
.filter-section{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:16px;margin-bottom:20px}
.filter-section h3{font-size:14px;margin-bottom:12px}
.filter-buttons{display:flex;flex-wrap:wrap;gap:8px;margin-bottom:12px}
.filter-btn{background:rgba(255,255,255,.08);border:1px solid var(--border);border-radius:6px;padding:6px 12px;font-size:12px;color:var(--muted);cursor:pointer;transition:all .2s}
.filter-btn:hover{background:rgba(88,166,255,.12);color:var(--blue);border-color:var(--blue)}
.filter-btn.active{background:rgba(88,166,255,.2);color:var(--blue);border-color:var(--blue)}
"""


def _item_html(item, important=False):
    cls = "item important" if important else "item"
    tags = f'<span class="tag tag-cat">{item["类型"]}</span>'
    if item["公司"] != "其他":
        tags += f'<span class="tag tag-co">{item["公司"]}</span>'
    if important:
        tags += '<span class="tag tag-imp">🔥 重要</span>'
    return (
        f'<div class="{cls}">'
        f'<div class="title">{item["资讯标题"]}</div>'
        f'<div class="meta2">{tags}</div>'
        f'<div class="points">{item["重点内容"]}</div>'
        f'<div style="margin-top:8px"><a href="{item["信息链接"]}" target="_blank">查看原文 →</a></div>'
        f"</div>"
    )


def generate_html_report(items, date_str, config):
    """生成HTML报告，按新布局展示"""
    imp = [n for n in items if n["是否重要"] == "是"]
    
    # 分类统计
    cats = {}
    for n in items:
        cats[n["类型"]] = cats.get(n["类型"], 0) + 1

    companies_set = {n["公司"] for n in items if n["公司"] != "其他"}

    stats_div = (
        f'<div class="stat"><div class="num">{len(items)}</div><div class="label">总资讯</div></div>'
        f'<div class="stat"><div class="num">{len(imp)}</div><div class="label">重要资讯</div></div>'
        f'<div class="stat"><div class="num">{len(cats)}</div><div class="label">覆盖类别</div></div>'
        f'<div class="stat"><div class="num">{len(companies_set)}</div><div class="label">涉及公司</div></div>'
    )

    # ===== 1. 重要资讯速览（列表形式）=====
    imp_section = ""
    if imp:
        imp_items_html = ""
        for i, n in enumerate(imp[:15], 1):
            imp_items_html += f'''
            <div class="important-item">
                <div class="marker">🔥</div>
                <div class="content">
                    <div class="title">{i}. {n["资讯标题"]}</div>
                    <div class="meta">
                        <span class="tag tag-cat">{n["类型"]}</span>
                        {f'<span class="tag tag-co">{n["公司"]}</span>' if n["公司"] != "其他" else ""}
                        <span style="margin-left:8px">{n["时间"]}</span>
                    </div>
                    <a href="{n['信息链接']}" target="_blank" class="link">查看原文 →</a>
                </div>
            </div>'''
        
        imp_section = f'''
        <div class="important-list">
            <h2>🔥 重要资讯速览 <span style="font-size:14px;color:var(--muted);margin-left:auto">{len(imp)}条</span></h2>
            {imp_items_html}
        </div>'''

    # ===== 2. 重点领域横向展示 =====
    # 定义重点领域（与smart_category函数中的规则匹配）
    key_domains = [
        {"name": "🎙️ 语音相关", "categories": ["语音技术"], "color": "#22c55e"},
        {"name": "👁️ 视觉相关", "categories": ["视觉相关", "视频生成", "图像生成"], "color": "#f59e0b"},
        {"name": "🎵 音频/音乐", "categories": ["音频/音乐"], "color": "#ec4899"},
        {"name": "🧠 大模型", "categories": ["大模型"], "color": "#3b82f6"},
        {"name": "🛡️ AI安全", "categories": ["AI安全"], "color": "#ef4444"},
        {"name": "📊 评测榜单", "categories": ["评测榜单"], "color": "#8b5cf6"},
    ]
    
    domain_items_map = {}
    used_for_domain = set()
    
    for domain in key_domains:
        domain_items = []
        for i, n in enumerate(items):
            if i in used_for_domain:
                continue
            if n["类型"] in domain["categories"]:
                domain_items.append(n)
                used_for_domain.add(i)
        if domain_items:
            domain_items_map[domain["name"]] = {
                "items": domain_items,
                "color": domain["color"]
            }
    
    domain_html = ""
    if domain_items_map:
        domain_cards = ""
        for domain_name, domain_data in domain_items_map.items():
            items_html = ""
            for n in domain_data["items"][:6]:
                items_html += f'''
                <div class="item">
                    <div class="title">{n["资讯标题"]}</div>
                    <div class="meta2">
                        <span class="tag tag-co">{n["公司"]}</span>
                        <span style="font-size:11px;color:var(--muted)">{n["时间"]}</span>
                    </div>
                    <div style="margin-top:4px"><a href="{n['信息链接']}" target="_blank" style="font-size:11px;color:var(--blue)">查看</a></div>
                </div>'''
            
            domain_cards += f'''
            <div class="domain-card" style="border-left:3px solid {domain_data['color']}">
                <h3 style="color:{domain_data['color']}">{domain_name} <span class="count">{len(domain_data['items'])}条</span></h3>
                <div class="items">{items_html}</div>
            </div>'''
        
        domain_html = f'''
        <section>
            <h2>🎯 重点领域</h2>
            <div class="domain-grid">{domain_cards}</div>
        </section>'''

    # ===== 3. 大厂动态（国外+国内）=====
    foreign_companies = ["OpenAI", "Google", "Microsoft", "Anthropic", "Meta", "xAI", "Stability AI"]
    domestic_companies = ["字节跳动", "阿里巴巴", "腾讯", "百度", "科大讯飞", "快手", "智谱AI", "月之暗面", "MiniMax", "DeepSeek"]
    
    def get_company_news(company_list, exclude_indices):
        """获取指定公司的资讯"""
        company_map = {}
        for company in company_list:
            company_items = []
            for i, n in enumerate(items):
                if i in exclude_indices:
                    continue
                if n["公司"] == company:
                    company_items.append(n)
                    exclude_indices.add(i)
            if company_items:
                company_map[company] = company_items
        return company_map
    
    used_for_company = set()
    foreign_news = get_company_news(foreign_companies, used_for_company)
    domestic_news = get_company_news(domestic_companies, used_for_company)
    
    company_html = ""
    if foreign_news or domestic_news:
        company_groups = ""
        
        # 国外大厂
        if foreign_news:
            foreign_grid = ""
            for company, news_list in foreign_news.items():
                news_items = ""
                for n in news_list[:3]:
                    news_items += f'''
                    <div class="item">
                        <div class="title">{n["资讯标题"]}</div>
                        <div class="meta2">
                            <span class="tag tag-cat">{n["类型"]}</span>
                        </div>
                        <a href="{n['信息链接']}" target="_blank" style="font-size:11px;color:var(--blue)">查看</a>
                    </div>'''
                foreign_grid += f'''
                <div class="company-item">
                    <div class="company-name">{company}</div>
                    {news_items}
                </div>'''
            
            company_groups += f'''
            <div class="company-group">
                <h3>🌍 国外：OpenAI - Google - Microsoft</h3>
                <div class="company-grid">{foreign_grid}</div>
            </div>'''
        
        # 国内大厂
        if domestic_news:
            domestic_grid = ""
            for company, news_list in domestic_news.items():
                news_items = ""
                for n in news_list[:3]:
                    news_items += f'''
                    <div class="item">
                        <div class="title">{n["资讯标题"]}</div>
                        <div class="meta2">
                            <span class="tag tag-cat">{n["类型"]}</span>
                        </div>
                        <a href="{n['信息链接']}" target="_blank" style="font-size:11px;color:var(--blue)">查看</a>
                    </div>'''
                domestic_grid += f'''
                <div class="company-item">
                    <div class="company-name">{company}</div>
                    {news_items}
                </div>'''
            
            company_groups += f'''
            <div class="company-group">
                <h3>🇨🇳 国内：字节 - 阿里 - MiniMax</h3>
                <div class="company-grid">{domestic_grid}</div>
            </div>'''
        
        company_html = f'''
        <section class="company-section">
            <h2>🏢 大厂动态</h2>
            {company_groups}
        </section>'''

    # ===== 4. AI提效资讯 =====
    efficiency_keywords = ["提效", "效率", "工具", "助手", "自动化", "workflow", "效率", "优化", "加速", "提速"]
    efficiency_items = []
    used_for_efficiency = set()
    
    for i, n in enumerate(items):
        if i in used_for_efficiency or i in used_for_domain or i in used_for_company:
            continue
        text = (n["资讯标题"] + " " + n["重点内容"]).lower()
        if any(kw in text for kw in efficiency_keywords):
            efficiency_items.append(n)
            used_for_efficiency.add(i)
    
    efficiency_html = ""
    if efficiency_items:
        eff_items = ""
        for n in efficiency_items[:8]:
            eff_items += f'''
            <div class="item">
                <div class="title">{n["资讯标题"]}</div>
                <div class="meta2">
                    <span class="tag tag-cat">{n["类型"]}</span>
                    <span class="tag tag-co">{n["公司"]}</span>
                </div>
                <div class="points">{n["重点内容"]}</div>
                <a href="{n['信息链接']}" target="_blank">查看原文 →</a>
            </div>'''
        
        efficiency_html = f'''
        <section class="efficiency-section">
            <h2>⚡ AI提效资讯 <span style="font-size:14px;color:var(--muted);margin-left:auto">{len(efficiency_items)}条</span></h2>
            {eff_items}
        </section>'''

    # ===== 5. 其他资讯（分类展示+筛选）=====
    # 收集所有未使用的资讯
    all_used = used_for_domain | used_for_company | used_for_efficiency
    other_items = []
    for i, n in enumerate(items):
        if i not in all_used:
            other_items.append(n)
    
    # 按类别分组
    other_grouped = {}
    for n in other_items:
        other_grouped.setdefault(n["类型"], []).append(n)
    
    # 收集所有类别用于筛选
    all_categories = sorted(other_grouped.keys())
    
    other_html = ""
    if other_grouped:
        # 生成筛选按钮
        filter_buttons = '<button class="filter-btn active" data-category="all">全部</button>'
        for cat in all_categories:
            count = len(other_grouped[cat])
            filter_buttons += f'<button class="filter-btn" data-category="{cat}">{cat} ({count})</button>'
        
        # 生成分类内容
        category_content = ""
        for cat in sorted(other_grouped.keys(), key=lambda c: len(other_grouped[c]), reverse=True):
            cat_items = ""
            for n in other_grouped[cat]:
                cat_items += f'''
                <div class="item" data-category="{cat}">
                    <div class="title">{n["资讯标题"]}</div>
                    <div class="meta2">
                        <span class="tag tag-cat">{n["类型"]}</span>
                        {f'<span class="tag tag-co">{n["公司"]}</span>' if n["公司"] != "其他" else ""}
                    </div>
                    <div class="points">{n["重点内容"]}</div>
                    <a href="{n['信息链接']}" target="_blank">查看原文 →</a>
                </div>'''
            category_content += f'''
            <div class="category-group" data-category="{cat}">
                <h3 class="other-cat">{cat}（{len(other_grouped[cat])}条）</h3>
                {cat_items}
            </div>'''
        
        other_html = f'''
        <section>
            <h2>📋 其他资讯 <span style="font-size:14px;color:var(--muted);margin-left:auto">{len(other_items)}条</span></h2>
            <div class="filter-section">
                <h3>筛选：</h3>
                <div class="filter-buttons">{filter_buttons}</div>
            </div>
            <div id="other-content">{category_content}</div>
        </section>'''

    # 分类统计条
    cat_stats = " | ".join(f"{c}: {v}条" for c, v in sorted(cats.items(), key=lambda x: -x[1]))

    html = f"""<!DOCTYPE html><html lang="zh"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>AI资讯日报 - {date_str}</title>
<style>{_CSS}</style>
<script>
document.addEventListener('DOMContentLoaded', function() {{
    // 筛选功能
    const filterBtns = document.querySelectorAll('.filter-btn');
    const categoryGroups = document.querySelectorAll('.category-group');
    
    filterBtns.forEach(btn => {{
        btn.addEventListener('click', function() {{
            // 移除所有激活状态
            filterBtns.forEach(b => b.classList.remove('active'));
            this.classList.add('active');
            
            const category = this.dataset.category;
            
            categoryGroups.forEach(group => {{
                if (category === 'all' || group.dataset.category === category) {{
                    group.style.display = 'block';
                }} else {{
                    group.style.display = 'none';
                }}
            }});
        }});
    }});
}});
</script>
</head>
<body><div class="header"><h1>👁️ AI行业资讯日报</h1><div class="meta">{date_str} | AI News Collector v3.0</div></div>
<div class="stats">{stats_div}</div>
{imp_section}
{domain_html}
{company_html}
{efficiency_html}
{other_html}
<div class="footer"><p>{cat_stats}</p><p style="margin-top:8px">由 AI资讯收集器 v3.0 自动生成</p></div>
</body></html>"""
    return html


# ========== 核心采集（优化） ==========
def collect_news(config):
    tr_cfg = config.get("translation", {})
    dedup_cfg = config.get("deduplication", {})
    sources = config.get("sources", [])
    
    load_translation_cache(tr_cfg)

    all_news = []
    seen_titles = []
    seen_links = set()
    seen_content = set()
    filtered = 0
    trans_fail = 0
    url_dedup_count = 0
    title_dedup_count = 0
    content_dedup_count = 0

    # 去重配置
    title_threshold = dedup_cfg.get("title_similarity_threshold", 0.85)
    enable_url_dedup = dedup_cfg.get("enable_url_dedup", True)
    enable_content_dedup = dedup_cfg.get("enable_content_dedup", True)

    for src in sources:
        limit = src.get("limit", 15)
        logging.info(f"📡 {src['name']}...")
        
        # 抓取RSS，带超时保护
        items = fetch_rss(src["url"], limit, timeout=30)
        if not items:
            logging.warning(f"  ⚠️ 未获取到数据")
            continue
        
        logging.info(f"  获取 {len(items)} 条")

        src_kept = 0
        for item in items:
            if not contains_ai_keyword(item["title"], item["summary"]):
                filtered += 1
                continue

            src_kept += 1
            src_type = src.get("type", "")

            # 1. URL去重（标准化）
            if enable_url_dedup:
                link = item.get("link", "").strip()
                normalized_url = normalize_url(link)
                if normalized_url and normalized_url in seen_links:
                    url_dedup_count += 1
                    continue

            # 2. 内容指纹去重
            if enable_content_dedup:
                raw_summary = clean_text(item.get("summary", ""), 200)
                content_fp = re.sub(r'\s+', '', raw_summary[:80]).lower()
                if content_fp and content_fp in seen_content:
                    content_dedup_count += 1
                    continue

            # 翻译
            if "中文" in src_type:
                title_text = item["title"]
                summary_text = item["summary"]
            else:
                title_text = translate_text(item["title"], tr_cfg)
                summary_text = translate_text(item["summary"], tr_cfg)
                if title_text == item["title"] and _is_english(item["title"]):
                    trans_fail += 1

            title_clean = clean_text(title_text, 120)
            if not title_clean:
                continue

            # 3. 标题相似度去重
            if is_duplicate_title(title_clean, seen_titles, title_threshold):
                title_dedup_count += 1
                continue

            # 全部通过，记录去重标记
            seen_titles.append(title_clean)
            if enable_url_dedup and normalized_url:
                seen_links.add(normalized_url)
            if enable_content_dedup and content_fp:
                seen_content.add(content_fp)

            summary_clean = clean_text(summary_text, 400)
            key_points = generate_key_points(summary_clean)
            
            # 使用真实的发布时间
            pub_time = item.get("published_parsed", datetime.now())
            time_str = pub_time.strftime("%Y-%m-%d %H:%M") if pub_time else datetime.now().strftime("%Y-%m-%d %H:%M")

            all_news.append({
                "时间": time_str,
                "公司": extract_company(item["title"], item["summary"]),
                "资讯标题": title_clean,
                "重点内容": key_points or "暂无摘要",
                "类型": smart_category(item["title"], item["summary"]),
                "是否重要": "是" if is_important(item["title"], item["summary"]) else "否",
                "信息链接": item["link"],
            })

        logging.info(f"  → AI相关 {src_kept} 条")
        time.sleep(0.5)  # 限流

    save_translation_cache(tr_cfg)
    all_news.sort(key=lambda x: 0 if x["是否重要"] == "是" else 1)
    
    logging.info(f"📊 去重统计: URL去重{url_dedup_count}条, 标题去重{title_dedup_count}条, 内容去重{content_dedup_count}条")
    
    return all_news, filtered, trans_fail


# ========== 输出 ==========
def save_csv(items, filename):
    """输出CSV文件（带样式，使用openpyxl）"""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        # 如果没有openpyxl，使用简单的CSV格式
        with open(filename, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["时间", "公司", "资讯标题", "总结的重点", "信息来源网址链接", "信息类别", "是否重要"])
            for item in items:
                writer.writerow([item[k] for k in ["时间", "公司", "资讯标题", "重点内容", "信息链接", "类型", "是否重要"]])
        return

    # 使用openpyxl创建带样式的Excel文件
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI资讯"

    # 定义样式
    # 表头样式：深蓝色背景，白色加粗文字
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 数据行样式
    row_font = Font(name="微软雅黑", size=10)
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 重要标记样式
    important_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # 浅黄色背景

    # 交替行颜色（浅灰绿色）
    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # 写入表头
    headers = ["时间", "公司名称", "文章标题", "总结的重点", "信息来源网址链接", "信息类别", "是否重要"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    for row_num, item in enumerate(items, 2):
        # 时间列：居中对齐
        time_cell = ws.cell(row=row_num, column=1)
        time_cell.value = item["时间"]
        time_cell.font = row_font
        time_cell.alignment = Alignment(horizontal="center", vertical="center")
        time_cell.border = thin_border
        if row_num % 2 == 0:  # 交替行颜色
            time_cell.fill = alt_row_fill

        # 公司名称列
        company_cell = ws.cell(row=row_num, column=2)
        company_cell.value = item["公司"]
        company_cell.font = row_font
        company_cell.alignment = data_alignment
        company_cell.border = thin_border
        if row_num % 2 == 0:
            company_cell.fill = alt_row_fill

        # 文章标题列
        title_cell = ws.cell(row=row_num, column=3)
        title_cell.value = item["资讯标题"]
        title_cell.font = row_font
        title_cell.alignment = data_alignment
        title_cell.border = thin_border
        if row_num % 2 == 0:
            title_cell.fill = alt_row_fill

        # 总结的重点列（自动换行）
        content_cell = ws.cell(row=row_num, column=4)
        content_cell.value = item["重点内容"]
        content_cell.font = row_font
        content_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        content_cell.border = thin_border
        if row_num % 2 == 0:
            content_cell.fill = alt_row_fill

        # 信息来源网址链接列
        link_cell = ws.cell(row=row_num, column=5)
        link_cell.value = item["信息链接"]
        link_cell.font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")  # 蓝色下划线
        link_cell.alignment = data_alignment
        link_cell.border = thin_border
        if row_num % 2 == 0:
            link_cell.fill = alt_row_fill
        # 添加超链接
        if item["信息链接"] and item["信息链接"].startswith("http"):
            link_cell.hyperlink = item["信息链接"]

        # 信息类别列
        cat_cell = ws.cell(row=row_num, column=6)
        cat_cell.value = item["类型"]
        cat_cell.font = row_font
        cat_cell.alignment = data_alignment
        cat_cell.border = thin_border
        if row_num % 2 == 0:
            cat_cell.fill = alt_row_fill

        # 是否重要列：居中，"是"显示为红色加粗
        imp_cell = ws.cell(row=row_num, column=7)
        imp_cell.value = item["是否重要"]
        imp_cell.alignment = Alignment(horizontal="center", vertical="center")
        imp_cell.border = thin_border
        if item["是否重要"] == "是":
            imp_cell.font = Font(name="微软雅黑", size=10, bold=True, color="FF0000")
            imp_cell.fill = important_fill  # 黄色背景
        else:
            imp_cell.font = row_font
        if row_num % 2 == 0 and item["是否重要"] != "是":
            imp_cell.fill = alt_row_fill

    # 调整列宽
    ws.column_dimensions['A'].width = 15  # 时间
    ws.column_dimensions['B'].width = 15  # 公司名称
    ws.column_dimensions['C'].width = 30  # 文章标题
    ws.column_dimensions['D'].width = 50  # 总结的重点
    ws.column_dimensions['E'].width = 40  # 信息来源网址链接
    ws.column_dimensions['F'].width = 15  # 信息类别
    ws.column_dimensions['G'].width = 10  # 是否重要

    # 设置表头行高
    ws.row_dimensions[1].height = 25

    # 保存文件（.xlsx格式）
    xlsx_filename = filename.replace('.csv', '.xlsx')
    wb.save(xlsx_filename)
    print(f"✅ Excel文件已生成: {xlsx_filename}")

    # 同时生成CSV格式（兼容性）
    with open(filename, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["时间", "公司名称", "文章标题", "总结的重点", "信息来源网址链接", "信息类别", "是否重要"])
        for item in items:
            writer.writerow([item[k] for k in ["时间", "公司", "资讯标题", "重点内容", "信息链接", "类型", "是否重要"]])


def save_list(items, filename):
    """输出表格格式列表文件"""
    COL_TIME = 18
    COL_COMPANY = 12
    COL_CAT = 10
    COL_IMP = 6
    COL_TITLE = 50

    sep = "+" + "-" * (COL_TIME + 2) + "+" + "-" * (COL_COMPANY + 2) + "+" \
        + "-" * (COL_TITLE + 2) + "+" + "-" * 32 + "+" \
        + "-" * (COL_CAT + 2) + "+" + "-" * (COL_IMP + 2) + "+" + "-" * 52 + "+"

    def _trunc(s, w):
        return s[:w-1] + "…" if len(s) > w else s

    def _wrap(s, w):
        """按宽度折行"""
        lines = []
        while len(s) > w:
            lines.append(s[:w])
            s = s[w:]
        if s:
            lines.append(s)
        return lines or [""]

    with open(filename, "w", encoding="utf-8") as f:
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        f.write(f"{'='*130}\n")
        f.write(f"  AI行业资讯列表 - {now}    共 {len(items)} 条\n")
        f.write(f"{'='*130}\n\n")

        # 表头
        f.write(sep + "\n")
        f.write(f"| {'时间':^{COL_TIME}} | {'公司':^{COL_COMPANY}} | "
                f"{'资讯标题':^{COL_TITLE}} | {'重点信息':30} | "
                f"{'分类':^{COL_CAT}} | {'重点':^{COL_IMP}} | {'链接':50} |\n")
        f.write(sep + "\n")

        for item in items:
            time_str = _trunc(item["时间"], COL_TIME)
            company = _trunc(item["公司"], COL_COMPANY)
            title = _trunc(item["资讯标题"], COL_TITLE)
            cat = _trunc(item["类型"], COL_CAT)
            imp = "⭐" if item["是否重要"] == "是" else ""
            link = item["信息链接"]
            points = item["重点内容"].replace("\n", " | ") if item["重点内容"] else "暂无"

            # 重点信息折行
            point_lines = _wrap(points, 30)
            # 标题折行
            title_lines = _wrap(title, COL_TITLE)
            # 链接折行
            link_lines = _wrap(link, 50)

            max_lines = max(len(point_lines), len(title_lines), len(link_lines), 1)

            for i in range(max_lines):
                t = title_lines[i] if i < len(title_lines) else ""
                p = point_lines[i] if i < len(point_lines) else ""
                l = link_lines[i] if i < len(link_lines) else ""
                # 首行显示时间/公司/分类/重点
                if i == 0:
                    f.write(f"| {time_str:<{COL_TIME}} | {company:<{COL_COMPANY}} | "
                            f"{t:<{COL_TITLE}} | {p:<30} | "
                            f"{cat:<{COL_CAT}} | {imp:^{COL_IMP}} | {l:<50} |\n")
                else:
                    f.write(f"| {'':<{COL_TIME}} | {'':<{COL_COMPANY}} | "
                            f"{t:<{COL_TITLE}} | {p:<30} | "
                            f"{'':<{COL_CAT}} | {'':^{COL_IMP}} | {l:<50} |\n")

            f.write(sep + "\n")


def save_html(items, filename, config):
    date_str = datetime.now().strftime("%Y-%m-%d")
    html = generate_html_report(items, date_str, config)
    with open(filename, "w", encoding="utf-8") as f:
        f.write(html)


# ========== 入口 ==========
def main():
    # 预解析 --config
    cfg_path = None
    for i, arg in enumerate(sys.argv):
        if arg == "--config" and i + 1 < len(sys.argv):
            cfg_path = sys.argv[i + 1]
            break

    config = load_config(cfg_path)

    ap = argparse.ArgumentParser(description="AI行业资讯收集器 v3.0 - 优化版", formatter_class=argparse.RawDescriptionHelpFormatter, epilog="""
用法示例:
  python ai_news_collector_optimized.py              # 运行一次
  python ai_news_collector_optimized.py --schedule   # 按config间隔循环采集
  python ai_news_collector_optimized.py --schedule --every 2   # 每2小时采集一次
  python ai_news_collector_optimized.py --config my_config.yaml
  python ai_news_collector_optimized.py --add-source "https://example.com/rss" --source-name "示例"
""")
    ap.add_argument("--run-once", action="store_true", help="只运行一次（默认）")
    ap.add_argument("--schedule", action="store_true", help="循环定时采集")
    ap.add_argument("--every", type=float, help="定时采集间隔（小时），覆盖config")
    ap.add_argument("--config", type=str, help="指定配置文件路径")
    ap.add_argument("--add-source", type=str, help="添加RSS源URL")
    ap.add_argument("--source-name", type=str, help="RSS源名称（配合--add-source）")
    ap.add_argument("--source-type", type=str, default="自定义", help="RSS源类型")
    ap.add_argument("--source-limit", type=int, default=15, help="RSS源抓取条数限制")
    ap.add_argument("--quiet", action="store_true", help="静默模式，只输出错误")
    args = ap.parse_args()

    if args.add_source:
        new_source = {
            "url": args.add_source,
            "name": args.source_name or args.add_source[:40],
            "type": args.source_type,
            "limit": args.source_limit,
        }
        if HAS_YAML and cfg_path or find_config():
            fp = find_config(cfg_path)
            if fp:
                with open(fp, encoding="utf-8") as f:
                    cfg_data = yaml.safe_load(f)
                cfg_data.setdefault("sources", []).append(new_source)
                with open(fp, "w", encoding="utf-8") as f:
                    yaml.dump(cfg_data, f, allow_unicode=True, default_flow_style=False)
                print(f"✅ 已添加RSS源: {new_source['name']}")
                print(f"   URL: {args.add_source}")
                return
        print("❌ 未找到config.yaml，无法添加源")
        return

    level = logging.WARNING if args.quiet else logging.INFO
    logging.basicConfig(level=level, format="%(message)s", handlers=[logging.StreamHandler()])

    interval = args.every or config.get("schedule", {}).get("interval_hours", 6)
    out = config.get("output", {})
    output_dir = out.get("output_dir", "")
    
    # 生成时间戳
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if output_dir:
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        csv_file = str(Path(output_dir) / f"{out.get('csv', 'ai_news_report').split('.')[0]}_{timestamp}.csv")
        html_file = str(Path(output_dir) / f"{out.get('html', 'ai_news_report').split('.')[0]}_{timestamp}.html")
        list_file = str(Path(output_dir) / f"{out.get('list', 'ai_news_list').split('.')[0]}_{timestamp}.txt")
    else:
        csv_file = f"{out.get('csv', 'ai_news_report').split('.')[0]}_{timestamp}.csv"
        html_file = f"{out.get('html', 'ai_news_report').split('.')[0]}_{timestamp}.html"
        list_file = f"{out.get('list', 'ai_news_list').split('.')[0]}_{timestamp}.txt"

    def run_once():
        logging.info("=" * 60)
        logging.info(f"🤖 AI资讯收集器 v3.0 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        logging.info("=" * 60)
        items, filtered, trans_fail = collect_news(config)
        if items:
            save_csv(items, csv_file)
            save_list(items, list_file)
            max_html = out.get("html_max_items", 100)
            save_html(items[:max_html], html_file, config)
        logging.info("=" * 60)
        logging.info(f"✅ 收集完成！共 {len(items)} 条AI相关资讯")
        logging.info(f"📊 已过滤 {filtered} 条（不含AI关键词）")
        if config.get("translation", {}).get("enabled"):
            logging.info(f"📝 翻译失败 {trans_fail} 条（已保留原文）")
        if items:
            logging.info(f"📁 CSV: {csv_file}")
            logging.info(f"📄 列表: {list_file}")
            logging.info(f"🌐 HTML: {html_file}")
            cats = {}
            for n in items:
                cats[n["类型"]] = cats.get(n["类型"], 0) + 1
            logging.info("\n📋 分类统计:")
            for c, v in sorted(cats.items(), key=lambda x: -x[1]):
                logging.info(f"   {c}: {v}条")
            imp = [n for n in items if n["是否重要"] == "是"]
            if imp:
                logging.info(f"\n🔥 重要资讯 ({len(imp)}条):")
                for i, item in enumerate(imp[:10], 1):
                    logging.info(f"\n{i}. 【{item['类型']}】{item['资讯标题']}")
                    logging.info(f"   🏢 {item['公司']}")
                    if item["重点内容"]:
                        for line in item["重点内容"].split("\n"):
                            if line.strip():
                                logging.info(f"   {line}")
                    logging.info(f"   🔗 {item['信息链接']}")
        else:
            logging.info("   今日暂无AI相关资讯")
        logging.info("=" * 60)
        return len(items)

    if args.schedule:
        delay = config.get("schedule", {}).get("first_run_delay_minutes", 0)
        if delay > 0:
            logging.info(f"⏰ 首次运行延迟 {delay} 分钟...")
            time.sleep(delay * 60)
        
        logging.info(f"⏰ 定时模式启动，每 {interval} 小时采集一次")
        logging.info("   按 Ctrl+C 退出\n")
        while True:
            try:
                run_once()
                logging.info(f"\n⏳ 下次采集: {interval} 小时后\n")
                time.sleep(interval * 3600)
            except KeyboardInterrupt:
                logging.info("\n👋 已停止定时采集")
                break
    else:
        run_once()


if __name__ == "__main__":
    main()
